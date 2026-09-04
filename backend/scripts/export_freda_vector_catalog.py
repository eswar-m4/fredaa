from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[2]
BOTS_PATH = ROOT / "frontend" / "src" / "data" / "bots.json"
OUT_PATH = ROOT / "data" / "freda_vector_db_agent_catalog.xlsx"


WORKFLOW_SOLUTIONS = [
    {
        "solution_name": "Website Verification",
        "solution_group": "Verification",
        "agents": "Website Discovery; Metadata Scraping; Confidence Scoring",
        "primary_sources": "company_website;linkedin;news;other",
        "data_types": "company website metadata; contact metadata; business identity; confidence score",
        "description": "Validates a company by discovering likely websites, scraping metadata, and scoring confidence before approval.",
        "output_fields": "website;legal_name;phone;email;linkedin_url;hq_address;description;confidence_score;source_notes",
        "workflow_stage": "Discovery and verification",
    },
    {
        "solution_name": "Contact Enrichment",
        "solution_group": "Enrichment",
        "agents": "Contact Enrichment",
        "primary_sources": "company_website;linkedin;news;other",
        "data_types": "email;phone;contact page;person/company contact metadata",
        "description": "Finds and validates direct contact data from website and public web sources, then enriches the record.",
        "output_fields": "possible_email;possible_phone;linkedin_url;contact_page_url;source_url",
        "workflow_stage": "Contact enrichment",
    },
    {
        "solution_name": "SEC Enrichment",
        "solution_group": "Registry / Financial",
        "agents": "SEC EDGAR",
        "primary_sources": "sec",
        "data_types": "filing metadata; company filings; SEC identifiers; financial disclosures",
        "description": "Pulls SEC EDGAR-based company and filing metadata to verify legal identity and financial context.",
        "output_fields": "cik;ticker;company_name;filing_metadata;filing_url;sec_status",
        "workflow_stage": "SEC registry enrichment",
    },
    {
        "solution_name": "MCA Enrichment",
        "solution_group": "Registry / Compliance",
        "agents": "Registry Enrichment",
        "primary_sources": "mca",
        "data_types": "company registry data; director info; incorporation data; legal status",
        "description": "Uses MCA/registry sources to enrich business entity records with legal and compliance metadata.",
        "output_fields": "legal_name;incorporation_date;directors;company_status;registry_metadata",
        "workflow_stage": "Registry enrichment",
    },
    {
        "solution_name": "Data Refresh",
        "solution_group": "Refresh",
        "agents": "Data Refresh",
        "primary_sources": "company_website;linkedin;sec;mca;news;other",
        "data_types": "fresh company, contact, registry, and market data",
        "description": "Refreshes stale records with the latest publicly available source signals and metadata.",
        "output_fields": "updated_business_metadata;fresh_contact_data;registry_data;confidence_shift",
        "workflow_stage": "Refresh and revalidation",
    },
]


def load_bots() -> list[dict]:
    payload = json.loads(BOTS_PATH.read_text(encoding="utf-8"))
    bots = payload.get("bots", []) if isinstance(payload, dict) else []
    return [bot for bot in bots if isinstance(bot, dict)]


def source_type_for(bot: dict) -> str:
    category = str(bot.get("category") or "")
    bot_type = str(bot.get("type") or "")
    if "Registry" in category or "SEC" in category:
        return "Registry / SEC"
    if bot_type.lower().startswith("listing") or "listing" in bot_type.lower():
        return "Listing / Product"
    if bot.get("url"):
        return "Website / Public Source"
    return "Custom"


def build_agent_rows() -> list[dict]:
    rows = []
    for index, bot in enumerate(load_bots(), start=1):
        name = str(bot.get("name") or "Unnamed Agent").strip()
        url = str(bot.get("url") or "").strip()
        category = str(bot.get("category") or "Uncategorized").strip()
        data_type = str(bot.get("dataType") or bot.get("data_type") or "").strip()
        rows.append(
            {
                "agent_id": index,
                "agent_name": name,
                "project": str(bot.get("project") or "").strip(),
                "source_url": url,
                "source_type": source_type_for(bot),
                "category": category,
                "industry": str(bot.get("industry") or "").strip(),
                "country": str(bot.get("country") or "").strip(),
                "data_type": data_type,
                "info": str(bot.get("info") or "").strip(),
                "type": str(bot.get("type") or "").strip(),
                "complexity": str(bot.get("complexity") or "").strip(),
                "datapoints": int(bot.get("datapoints") or 0),
                "catalog_kind": bot.get("catalog_kind") or "built_in",
                "source_key": str(bot.get("source_key") or "").strip(),
                "solution_use_case": f"Collects {data_type or 'public'} data from {name} for {category} workflows.",
                "vector_db_chunk": (
                    f"Agent: {name}. Category: {category}. Source: {url}. "
                    f"Country: {bot.get('country') or 'Unknown'}. Industry: {bot.get('industry') or 'Unknown'}. "
                    f"Data Type: {data_type or 'Unknown'}. Information: {bot.get('info') or 'Not specified'}. "
                    f"Complexity: {bot.get('complexity') or 'Unknown'}. Datapoints: {bot.get('datapoints') or 0}."
                ),
            }
        )
    return rows


def apply_header_style(worksheet) -> None:
    header_fill = PatternFill(fillType="solid", fgColor="2F75B5")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        for cell in column_cells:
            try:
                value = cell.value or ""
                max_length = max(max_length, len(str(value)))
            except Exception:
                pass
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(38, max(12, max_length + 2))


def build_workbook() -> Workbook:
    workbook = Workbook()
    agent_rows = build_agent_rows()

    agents_ws = workbook.active
    agents_ws.title = "Agent_Catalog"
    agent_headers = [
        "agent_id", "agent_name", "project", "source_url", "source_type", "category", "industry",
        "country", "data_type", "info", "type", "complexity", "datapoints", "catalog_kind", "source_key",
        "solution_use_case", "vector_db_chunk",
    ]
    agents_ws.append(agent_headers)
    for row in agent_rows:
        agents_ws.append([row.get(header, "") for header in agent_headers])
    apply_header_style(agents_ws)

    solution_ws = workbook.create_sheet("Workflow_Solutions")
    solution_headers = [
        "solution_name", "solution_group", "agents", "primary_sources", "data_types", "description",
        "output_fields", "workflow_stage",
    ]
    solution_ws.append(solution_headers)
    for solution in WORKFLOW_SOLUTIONS:
        solution_ws.append([solution.get(header, "") for header in solution_headers])
    apply_header_style(solution_ws)

    source_ws = workbook.create_sheet("Source_Reference")
    source_headers = [
        "source_name", "source_url", "source_type", "category", "country", "industry", "data_type", "info",
        "complexity", "datapoints",
    ]
    source_ws.append(source_headers)
    seen = set()
    for row in agent_rows:
        key = (row["agent_name"], row["source_url"])
        if key in seen:
            continue
        seen.add(key)
        source_ws.append(
            [
                row.get("agent_name", ""),
                row.get("source_url", ""),
                row.get("source_type", ""),
                row.get("category", ""),
                row.get("country", ""),
                row.get("industry", ""),
                row.get("data_type", ""),
                row.get("info", ""),
                row.get("complexity", ""),
                row.get("datapoints", 0),
            ]
        )
    apply_header_style(source_ws)

    vector_ws = workbook.create_sheet("Vector_DB_Feed_Ready")
    vector_headers = [
        "entity_type", "name", "url", "category", "source_type", "country", "industry",
        "data_type", "description", "content_for_vector_db",
    ]
    vector_ws.append(vector_headers)
    for row in agent_rows:
        vector_ws.append(
            [
                "agent",
                row.get("agent_name", ""),
                row.get("source_url", ""),
                row.get("category", ""),
                row.get("source_type", ""),
                row.get("country", ""),
                row.get("industry", ""),
                row.get("data_type", ""),
                row.get("info", ""),
                row.get("vector_db_chunk", ""),
            ]
        )
    for solution in WORKFLOW_SOLUTIONS:
        vector_ws.append(
            [
                "solution",
                solution.get("solution_name", ""),
                "",
                solution.get("solution_group", ""),
                solution.get("primary_sources", ""),
                "",
                "",
                solution.get("data_types", ""),
                solution.get("description", ""),
                (
                    f"Solution: {solution.get('solution_name')}. Agents: {solution.get('agents')}. "
                    f"Sources: {solution.get('primary_sources')}. Data types: {solution.get('data_types')}. "
                    f"Workflow: {solution.get('workflow_stage')}."
                ),
            ]
        )
    apply_header_style(vector_ws)

    for ws in workbook.worksheets:
        autosize_columns(ws)

    return workbook


if __name__ == "__main__":
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(OUT_PATH)
    print(f"Workbook created: {OUT_PATH}")
    print(f"Agents: {len(build_agent_rows())}")
    print(f"Sheets: {[ws.title for ws in workbook.worksheets]}")
